import csv
import json
import os
import queue
import sys
import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Callable, Iterable, List

from openpyxl import load_workbook


def _row_is_empty(row: Iterable[object | None]) -> bool:
    return all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row)


def convert_path(
    input_path: Path,
    output_dir: Path,
    sheet: str,
    start_row: int,
    progress_cb: Callable[[float], None] | None = None,
    combie_duo: bool = True,
) -> List[Path]:
    """Convert a single Excel file or every Excel file in a folder to CSV outputs.

    The primary export targets the given ``sheet`` starting at ``start_row``.
    When ``combie_duo`` is True, an additional export of the "Aviti Manifest"
    sheet (starting at row 16) is produced for each workbook.
    """

    if not input_path.exists():
        raise FileNotFoundError(f"Input path not found: {input_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    work_items: List[Path]
    if input_path.is_dir():
        work_items = [p for p in sorted(input_path.glob("*.xlsx")) if not p.name.startswith("~$")]
    else:
        work_items = [input_path]

    if not work_items:
        raise FileNotFoundError("No .xlsx files found to convert.")

    outputs: List[Path] = []
    total_exports = len(work_items) * (2 if combie_duo else 1)
    completed = 0

    def notify_progress() -> None:
        if progress_cb and total_exports:
            pct = min(100.0, max(0.0, (completed / total_exports) * 100.0))
            progress_cb(pct)

    for workbook_path in work_items:
        wb = load_workbook(workbook_path, data_only=True)
        targets = [(sheet, start_row)]
        if combie_duo:
            targets.append(("Aviti Manifest", 16))

        for sheet_name, first_row in targets:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path.name}")

            ws = wb[sheet_name]
            out_file = output_dir / f"{workbook_path.stem}_{sheet_name.replace(' ', '_')}.csv"
            with out_file.open("w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                for row in ws.iter_rows(min_row=first_row, values_only=True):
                    if _row_is_empty(row):
                        break
                    writer.writerow([cell if cell is not None else "" for cell in row])

            outputs.append(out_file)
            completed += 1
            notify_progress()

        wb.close()

    if progress_cb:
        progress_cb(100.0)

    return outputs

class ConvertFrame(ttk.Frame):
    """Screen for converting data with file or folder workflows."""

    def __init__(self, parent: tk.Widget, controller) -> None:
        super().__init__(parent, padding=18)
        self.controller = controller
        self._settings_path = Path(__file__).resolve().parent / "settings.json"

        self.progress_var = tk.DoubleVar(value=0.0)
        self.output_path_var = tk.StringVar()
        self._worker_thread: threading.Thread | None = None
        self._event_queue: queue.Queue[tuple[str, float | str]] = queue.Queue()
        self._running = False

        self.input_file_var = tk.StringVar()
        self.input_folder_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.combie_duo_var = tk.BooleanVar(value=True)

        self._load_settings()
        self._build_ui()

    def _build_ui(self) -> None:
        # Header row with title and Home navigation.
        header = ttk.Frame(self)
        header.pack(fill=tk.X, pady=(10, 20), padx=6)

        ttk.Label(header, text="Convert Data", font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT)
        ttk.Button(
            header,
            text="Home",
            width=7,
            command=lambda: self.controller.show_frame("home"),
            style="Card.TButton",
        ).pack(side=tk.RIGHT)

        form = ttk.Frame(self)
        form.pack(fill=tk.BOTH, expand=True, padx=6)

        # Inputs for file, folder, and output destination.
        self._add_path_selector(form, "Input File", self.input_file_var, self._browse_file, row=0)
        self._add_path_selector(form, "Input Folder", self.input_folder_var, self._browse_directory, row=1)
        self._add_path_selector(form, "Output Folder", self.output_var, self._browse_directory, row=2)

        # Action buttons for file vs. folder conversion.
        action_row = ttk.Frame(form)
        action_row.grid(row=3, column=0, columnspan=3, pady=(24, 6), sticky="w")
        self.convert_file_btn = ttk.Button(
            action_row,
            text="Convert File",
            style="Primary.TButton",
            command=lambda: self._start_convert(mode="file"),
        )
        self.convert_file_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.convert_folder_btn = ttk.Button(
            action_row,
            text="Convert Folder",
            style="Primary.TButton",
            command=lambda: self._start_convert(mode="folder"),
        )
        self.convert_folder_btn.pack(side=tk.LEFT)

        # Options
        options_row = ttk.Frame(form)
        options_row.grid(row=4, column=0, columnspan=3, pady=(6, 12), sticky="w")
        ttk.Checkbutton(
            options_row,
            text="Combie duo (SampleImport + Aviti Manifest)",
            variable=self.combie_duo_var,
            onvalue=True,
            offvalue=False,
        ).pack(side=tk.LEFT)

        # Progress section.
        progress_row = ttk.Frame(form)
        progress_row.grid(row=5, column=0, columnspan=3, pady=(10, 10), sticky="we")
        form.columnconfigure(1, weight=1)
        ttk.Label(progress_row, text="Progress:").pack(side=tk.LEFT, padx=(0, 8))
        self.progress = ttk.Progressbar(progress_row, variable=self.progress_var, maximum=100)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress_pct = ttk.Label(progress_row, text="0%")
        self.progress_pct.pack(side=tk.LEFT, padx=(8, 0))

        # Output row with open-folder helper.
        result_row = ttk.Frame(form)
        result_row.grid(row=6, column=0, columnspan=3, pady=(20, 0), sticky="we")
        ttk.Label(result_row, text="Output:").pack(side=tk.LEFT, padx=(0, 8))
        self.output_path_display = ttk.Entry(result_row, textvariable=self.output_path_var, state="readonly")
        self.output_path_display.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(result_row, text="Open Output Folder", command=self._open_output_folder).pack(side=tk.LEFT, padx=(8, 0))

    def _add_path_selector(
        self,
        parent: ttk.Frame,
        label_text: str,
        variable: tk.StringVar,
        browse_callback,
        row: int,
    ) -> None:
        label = ttk.Label(parent, text=label_text)
        label.grid(row=row, column=0, sticky="w", pady=6, padx=(0, 8))

        entry = ttk.Entry(parent, textvariable=variable)
        entry.grid(row=row, column=1, sticky="we", pady=6)

        btn = ttk.Button(
            parent,
            text="Browse",
            width=7,
            command=lambda: browse_callback(variable),
            style="Card.TButton",
        )
        btn.grid(row=row, column=2, sticky="e", pady=6, padx=(8, 0))

    def _browse_file(self, target: tk.StringVar) -> None:
        path = filedialog.askopenfilename(title="Select file to convert")
        if path:
            target.set(path)

    def _browse_directory(self, target: tk.StringVar) -> None:
        path = filedialog.askdirectory(title="Select folder")
        if path:
            target.set(path)

    def _start_convert(self, mode: str) -> None:
        # Kick off background conversion (file or folder) to keep UI responsive.
        if self._running:
            messagebox.showinfo("In progress", "Please wait until the current convert finishes.")
            return

        output = self.output_var.get().strip()
        input_path = self.input_file_var.get().strip() if mode == "file" else self.input_folder_var.get().strip()

        if not (input_path and output):
            messagebox.showwarning("Missing input", "Please provide input and output paths before starting.")
            return

        input_path_obj = Path(input_path)
        output_path_obj = Path(output)

        self._save_settings({
            "input_file": self.input_file_var.get().strip(),
            "input_folder": self.input_folder_var.get().strip(),
            "output": output,
            "combie_duo": self.combie_duo_var.get(),
        })

        self.progress_var.set(0)
        self.progress_pct.config(text="0%")
        self.output_path_var.set("")
        self._running = True
        self.convert_file_btn.state(["disabled"])
        self.convert_folder_btn.state(["disabled"])

        self._worker_thread = threading.Thread(
            target=self._convert_worker,
            args=(mode, input_path_obj, output_path_obj),
            daemon=True,
        )
        self._worker_thread.start()
        self.after(100, self._poll_events)

    def _convert_worker(self, mode: str, input_path: Path, output_path: Path) -> None:
        try:
            # Default sheet/start-row for combie outputs
            sheet = "SampleImport"
            start_row = 24
            combie_duo = self.combie_duo_var.get()

            def progress_cb(pct: float) -> None:
                self._event_queue.put(("progress", pct))

            convert_path(
                input_path=input_path,
                output_dir=output_path,
                sheet=sheet,
                start_row=start_row,
                progress_cb=progress_cb,
                combie_duo=combie_duo,
            )

            self._event_queue.put(("done", str(output_path), mode))
        except Exception as exc:  # pragma: no cover - defensive UI error handling
            self._event_queue.put(("error", str(exc)))

    def _poll_events(self) -> None:
        # Drain worker events and update UI safely from the main thread.
        try:
            while True:
                event_payload = self._event_queue.get_nowait()
                event = event_payload[0]
                if event == "progress":
                    payload = event_payload[1]
                    self.progress_var.set(float(payload))
                    self.progress_pct.config(text=f"{float(payload):.0f}%")
                elif event == "done":
                    _, output, mode = event_payload
                    self.progress_var.set(100)
                    self.progress_pct.config(text="100%")
                    self.output_path_var.set(str(output))
                    self._on_complete(success=True, message=f"Convert {mode} completed.")
                elif event == "error":
                    _, err = event_payload
                    self._on_complete(success=False, message=str(err))
        except queue.Empty:
            pass

        if self._running:
            self.after(100, self._poll_events)

    def _on_complete(self, success: bool, message: str) -> None:
        self._running = False
        self.convert_file_btn.state(["!disabled"])
        self.convert_folder_btn.state(["!disabled"])
        if success:
            messagebox.showinfo("Success", message)
        else:
            messagebox.showerror("Error", message)

    def _open_output_folder(self) -> None:
        path = self.output_path_var.get()
        if not path:
            messagebox.showinfo("No output", "Run convert to generate output first.")
            return

        if os.path.isdir(path):
            self._open_path(path)
        else:
            messagebox.showinfo("Not found", "The output folder does not exist.")

    @staticmethod
    def _open_path(path: str) -> None:
        try:
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                os.system(f"open '{path}'")
            else:
                os.system(f"xdg-open '{path}'")
        except Exception as exc:  # pragma: no cover - defensive UI error handling
            messagebox.showerror("Unable to open", str(exc))

    def _load_settings(self) -> None:
        try:
            if not self._settings_path.exists():
                return
            data = json.loads(self._settings_path.read_text(encoding="utf-8") or "{}")
            convert = data.get("convert", {})
            self.input_file_var.set(convert.get("input_file", ""))
            self.input_folder_var.set(convert.get("input_folder", ""))
            self.output_var.set(convert.get("output", ""))
            self.combie_duo_var.set(convert.get("combie_duo", True))
        except Exception:
            pass

    def _save_settings(self, convert_data: dict) -> None:
        try:
            data = {}
            if self._settings_path.exists():
                data = json.loads(self._settings_path.read_text(encoding="utf-8") or "{}")
            data["convert"] = convert_data
            self._settings_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
