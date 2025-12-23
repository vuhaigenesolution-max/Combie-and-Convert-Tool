from __future__ import annotations
from pathlib import Path
import re
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Filename pattern: metadata_RUNNAME_YYYYMMDD.xlsx or metadata_RUNNAME_YYYYMMDD_xxx.xlsx
FILENAME_REGEX = re.compile(r"^metadata_(?P<run>[A-Za-z0-9_-]+)_(?P<date>20\d{6})(?:_.*)?\.xlsx$", re.IGNORECASE)

# Source sheet & positions (1-based Excel rows/cols)
SOURCE_SHEET = "Sample"
IMPORT_SHEET = "SampleImport"
AVITI_SHEET = "Aviti Manifest"
AVITI_TEST_SHEET = "Aviti Manifest TEST"
SAMPLE_START_ROW = 22  # data starts on Sample sheet
IMPORT_START_ROW = 24  # data starts on SampleImport sheet
AVITI_START_ROW = 16  # data starts on Aviti Manifest sheet
AVITI_TEST_START_ROW = 24  # data starts on Aviti Manifest TEST sheet


def _sampleimport_col_k_formula(sample_row: int) -> str:
    # Build nested IF per business rules, referencing Sample sheet row.
    return (
        f"=IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"E\",LEFT(Sample!A{sample_row},1)=\"H\",LEFT(Sample!A{sample_row},1)=\"T\","  # TS1
        f"LEFT(Sample!A{sample_row},1)=\"B\",LEFT(Sample!A{sample_row},2)=\"ID\"),"
        f"OR(LEFT(Sample!C{sample_row},2)=\"JI\",LEFT(Sample!C{sample_row},1)=\"I\")),\"TS1\"," 
        f"IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"E\",LEFT(Sample!A{sample_row},1)=\"H\",LEFT(Sample!A{sample_row},1)=\"T\","  # TS95
        f"LEFT(Sample!A{sample_row},1)=\"B\",LEFT(Sample!A{sample_row},2)=\"ID\"),"
        f"OR(LEFT(Sample!C{sample_row},2)=\"JX\",LEFT(Sample!C{sample_row},2)=\"JW\",LEFT(Sample!C{sample_row},1)=\"X\")),\"TS95\"," 
        f"IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"E\",LEFT(Sample!A{sample_row},1)=\"H\",LEFT(Sample!A{sample_row},1)=\"T\","  # TS3
        f"LEFT(Sample!A{sample_row},1)=\"B\",LEFT(Sample!A{sample_row},2)=\"ID\"),"
        f"OR(LEFT(Sample!C{sample_row},2)=\"JN\",LEFT(Sample!C{sample_row},1)=\"N\")),\"TS3\"," 
        f"IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"E\",LEFT(Sample!A{sample_row},1)=\"H\",LEFT(Sample!A{sample_row},1)=\"T\","  # TS24
        f"LEFT(Sample!A{sample_row},1)=\"B\",LEFT(Sample!A{sample_row},2)=\"ID\"),"
        f"OR(LEFT(Sample!C{sample_row},2)=\"JA\",LEFT(Sample!C{sample_row},2)=\"AA\",LEFT(Sample!C{sample_row},2)=\"JS\",LEFT(Sample!C{sample_row},2)=\"SA\")),\"TS24\"," 
        f"IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"T\",LEFT(Sample!A{sample_row},1)=\"B\"),LEFT(Sample!C{sample_row},2)=\"AS\"),\"TSPRO\","  # TSPRO (AS)
        f"IF(AND(OR(LEFT(Sample!A{sample_row},1)=\"T\",LEFT(Sample!A{sample_row},1)=\"B\",LEFT(Sample!A{sample_row},1)=\"E\",LEFT(Sample!A{sample_row},1)=\"H\"),"
        f"LEFT(Sample!C{sample_row},4)=\"SERA\"),\"TSPRO\","  # TSPRO (SERA)
        f"IF(LEFT(Sample!A{sample_row},2)=\"CR\",\"CARRIER9\",IF(LEFT(Sample!A{sample_row},4)=\"DEL3\",\"NIPTDEL3\",Sample!A{sample_row}))))))))"  # default
    )

# Column mapping (source → template)
# Excel letters to letters; we’ll convert to column indices when writing.
COLUMN_MAP = {
    "A": "A",  # expNum
    "B": "B",  # sampleOrder
    "C": "C",  # LABCODE
    "D": "D",  # SeqType
    "E": "E",  # Harvest kit
    "F": "F",  # Harvest By
    "G": "G",  # Harvest Date
    "H": "H",  # Library By
    "I": "I",  # Library Date
    "J": "J",  # Species
    "K": "K",  # i7 index
    "L": "L",  # i5 index
    "M": "M",  # LibraryConc
    "N": "N",  # LibraryAmp
    "O": "O",  # Library Protocol
    "P": "P",  # LRMtemplate
    "Q": "Q",  # passedQC
    "R": "R",  # LANE
    "S": "S",  # TE 0.1X
    "T": "T",  # Primers
    "U": "U",  # Notes
}
CHECK_PRIMERS_COL = "V"
CHECK_LABCODES_COL = "W"


def _validate_inputs(source_folder: str, output_folder: str, template_file: str) -> Tuple[Path, Path, Path]:
    src = Path(source_folder)
    out = Path(output_folder)
    tpl = Path(template_file)

    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"Source folder not found: {src}")
    if not tpl.exists() or not tpl.is_file():
        raise FileNotFoundError(f"Template file not found: {tpl}")

    out.mkdir(parents=True, exist_ok=True)
    return src, out, tpl


def _iter_metadata_files(src: Path) -> List[Path]:
    files: List[Path] = []
    for path in src.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue  # skip Excel temp files
        if FILENAME_REGEX.match(path.name):
            files.append(path)
    return files


def _group_by_run_date(files: List[Path]) -> Dict[Tuple[str, str], List[Path]]:
    groups: Dict[Tuple[str, str], List[Path]] = {}
    for f in files:
        m = FILENAME_REGEX.match(f.name)
        if not m:
            continue
        key = (m.group("run"), m.group("date"))
        groups.setdefault(key, []).append(f)
    return groups


def _read_sample_rows(path: Path) -> List[Dict[str, object]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    if SOURCE_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws: Worksheet = wb[SOURCE_SHEET]

    rows: List[Dict[str, object]] = []
    for row in ws.iter_rows(min_row=SAMPLE_START_ROW):
        # stop when LABCODE (col C) is empty; this matches “đến hết labcode thì thôi”
        if len(row) >= 3:
            labcode_val = row[2].value  # column C
            if labcode_val is None or (isinstance(labcode_val, str) and labcode_val.strip() == ""):
                break
        record: Dict[str, object] = {}
        for col_letter in COLUMN_MAP.keys():
            col_idx = _col_letter_to_index(col_letter) - 1  # zero-based index in row tuple
            if col_idx < len(row):
                record[col_letter] = row[col_idx].value
            else:
                record[col_letter] = None
        rows.append(record)
    wb.close()
    return rows


def _col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (A=1) to index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _index_from_letter(letter: str) -> int:
    """Excel letter to 1-based column number."""
    return _col_letter_to_index(letter)


def _compute_duplicates(rows: List[Dict[str, object]]) -> Tuple[set, set]:
    primers_count: Dict[str, int] = {}
    labcode_count: Dict[str, int] = {}

    def _bump(counter: Dict[str, int], val: object):
        if val is None:
            return
        s = str(val).strip()
        if s == "":
            return
        counter[s] = counter.get(s, 0) + 1

    for r in rows:
        _bump(primers_count, r.get("T"))  # Primers
        _bump(labcode_count, r.get("C"))  # LABCODE

    dup_primers = {k for k, v in primers_count.items() if v > 1}
    dup_labcodes = {k for k, v in labcode_count.items() if v > 1}
    return dup_primers, dup_labcodes


def _write_group(run: str, date: str, rows: List[Dict[str, object]], template_file: Path, output_dir: Path) -> Path:
    # Fresh load of template to avoid touching the original
    wb = load_workbook(template_file, data_only=False)
    if SOURCE_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Template missing sheet: {SOURCE_SHEET}")
    if IMPORT_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Template missing sheet: {IMPORT_SHEET}")
    if AVITI_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Template missing sheet: {AVITI_SHEET}")
    if AVITI_TEST_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Template missing sheet: {AVITI_TEST_SHEET}")

    sample_ws: Worksheet = wb[SOURCE_SHEET]
    import_ws: Worksheet = wb[IMPORT_SHEET]
    aviti_ws: Worksheet = wb[AVITI_SHEET]
    aviti_test_ws: Worksheet = wb[AVITI_TEST_SHEET]

    dup_primers, dup_labcodes = _compute_duplicates(rows)

    sample_start_row = SAMPLE_START_ROW
    import_start_row = IMPORT_START_ROW
    aviti_start_row = AVITI_START_ROW
    aviti_test_start_row = AVITI_TEST_START_ROW

    for offset, record in enumerate(rows):
        sample_row = sample_start_row + offset
        import_row = import_start_row + offset
        aviti_row = aviti_start_row + offset
        aviti_test_row = aviti_test_start_row + offset
        # Write mapped columns on Sample sheet
        for src_col, dst_col in COLUMN_MAP.items():
            val = record.get(src_col)
            cell = sample_ws.cell(row=sample_row, column=_index_from_letter(dst_col))
            cell.value = val
            if dst_col == "I":
                cell.number_format = "DD/MM/YYYY"

        # Fill formula in col K to lookup i7 index from Primers (col T)
        sample_ws.cell(row=sample_row, column=_index_from_letter("K")).value = (
            f"=VLOOKUP(T{sample_row},'Index Sets'!$A$2:$C$4000,2,FALSE)"
        )

        # SampleImport formulas
        import_ws.cell(row=import_row, column=_index_from_letter("A")).value = (
            f"=Sample!B{sample_row}&\"-\"&Sample!C{sample_row}"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("B")).value = (
            f"=Sample!B{sample_row}&\"-\"&Sample!C{sample_row}"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("C")).value = (
            f"=Sample!A{sample_row}"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("F")).value = (
            f"=Sample!K{sample_row}"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("G")).value = (
            f"=VLOOKUP(F{import_row},'Index Sequence'!$A$2:$B$10000,2,FALSE)"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("H")).value = (
            f"=Sample!L{sample_row}"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("I")).value = (
            f"=VLOOKUP(H{import_row},'Index Sequence'!$A$2:$B$10000,2,FALSE)"
        )
        import_ws.cell(row=import_row, column=_index_from_letter("K")).value = _sampleimport_col_k_formula(sample_row)

        # Aviti Manifest formulas
        aviti_ws.cell(row=aviti_row, column=_index_from_letter("A")).value = (
            f"=SampleImport!A{import_row}"
        )
        aviti_ws.cell(row=aviti_row, column=_index_from_letter("B")).value = (
            f"=SampleImport!G{import_row}"
        )
        # Reverse string in SampleImport col I (positions 1..30) into Aviti col C
        aviti_ws.cell(row=aviti_row, column=_index_from_letter("C")).value = (
            f"=MID(I{aviti_row},30,1)&MID(I{aviti_row},29,1)&MID(I{aviti_row},28,1)&MID(I{aviti_row},27,1)&"
            f"MID(I{aviti_row},26,1)&MID(I{aviti_row},25,1)&MID(I{aviti_row},24,1)&MID(I{aviti_row},23,1)&"
            f"MID(I{aviti_row},22,1)&MID(I{aviti_row},21,1)&MID(I{aviti_row},20,1)&MID(I{aviti_row},19,1)&"
            f"MID(I{aviti_row},18,1)&MID(I{aviti_row},17,1)&MID(I{aviti_row},16,1)&MID(I{aviti_row},15,1)&"
            f"MID(I{aviti_row},14,1)&MID(I{aviti_row},13,1)&MID(I{aviti_row},12,1)&MID(I{aviti_row},11,1)&"
            f"MID(I{aviti_row},10,1)&MID(I{aviti_row},9,1)&MID(I{aviti_row},8,1)&MID(I{aviti_row},7,1)&"
            f"MID(I{aviti_row},6,1)&MID(I{aviti_row},5,1)&MID(I{aviti_row},4,1)&MID(I{aviti_row},3,1)&"
            f"MID(I{aviti_row},2,1)&MID(I{aviti_row},1,1)"
        )
        aviti_ws.cell(row=aviti_row, column=_index_from_letter("D")).value = (
            f"=SampleImport!K{import_row}"
        )
        aviti_ws.cell(row=aviti_row, column=_index_from_letter("I")).value = (
            f"=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SampleImport!I{import_row},\"A\",1),\"C\",2),\"G\",3),\"T\",4),1,\"T\"),2,\"G\"),3,\"C\"),4,\"A\")"
        )

        # Aviti Manifest TEST formulas (start row 24)
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("A")).value = (
            f"=Sample!B{sample_row}&\"-\"&Sample!C{sample_row}"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("B")).value = (
            f"=Sample!B{sample_row}&\"-\"&Sample!C{sample_row}"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("C")).value = (
            f"=Sample!A{sample_row}"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("F")).value = (
            f"=Sample!K{sample_row}"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("G")).value = (
            f"=VLOOKUP(F{aviti_test_row},'Index Sequence'!$A$2:$B$9808,2,FALSE)"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("H")).value = (
            f"=Sample!L{sample_row}"
        )
        # O column depends on SampleImport col I
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("O")).value = (
            f"=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SampleImport!I{import_row},\"A\",1),\"C\",2),\"G\",3),\"T\",4),1,\"T\"),2,\"G\"),3,\"C\"),4,\"A\")"
        )
        aviti_test_ws.cell(row=aviti_test_row, column=_index_from_letter("I")).value = (
            f"=MID(O{aviti_test_row},30,1)&MID(O{aviti_test_row},29,1)&MID(O{aviti_test_row},28,1)&MID(O{aviti_test_row},27,1)&"
            f"MID(O{aviti_test_row},26,1)&MID(O{aviti_test_row},25,1)&MID(O{aviti_test_row},24,1)&MID(O{aviti_test_row},23,1)&"
            f"MID(O{aviti_test_row},22,1)&MID(O{aviti_test_row},21,1)&MID(O{aviti_test_row},20,1)&MID(O{aviti_test_row},19,1)&"
            f"MID(O{aviti_test_row},18,1)&MID(O{aviti_test_row},17,1)&MID(O{aviti_test_row},16,1)&MID(O{aviti_test_row},15,1)&"
            f"MID(O{aviti_test_row},14,1)&MID(O{aviti_test_row},13,1)&MID(O{aviti_test_row},12,1)&MID(O{aviti_test_row},11,1)&"
            f"MID(O{aviti_test_row},10,1)&MID(O{aviti_test_row},9,1)&MID(O{aviti_test_row},8,1)&MID(O{aviti_test_row},7,1)&"
            f"MID(O{aviti_test_row},6,1)&MID(O{aviti_test_row},5,1)&MID(O{aviti_test_row},4,1)&MID(O{aviti_test_row},3,1)&"
            f"MID(O{aviti_test_row},2,1)&MID(O{aviti_test_row},1,1)"
        )

        primers_val = record.get("T")
        labcode_val = record.get("C")
        # Clear markers; no Y/N output requested
        sample_ws.cell(row=sample_row, column=_index_from_letter(CHECK_PRIMERS_COL)).value = ""
        sample_ws.cell(row=sample_row, column=_index_from_letter(CHECK_LABCODES_COL)).value = ""

    out_path = output_dir / f"{run}_{date}.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path


def run_export(source_folder: str, output_folder: str, template_file: str) -> List[Path]:
    """
    Entry point for the combie UI.
    - source_folder: folder containing metadata_*.xlsx files
    - output_folder: destination folder (created if missing)
    - template_file: Excel template to copy & fill
    Returns list of generated file paths.
    """
    src, out, tpl = _validate_inputs(source_folder, output_folder, template_file)

    files = _iter_metadata_files(src)
    if not files:
        raise FileNotFoundError("No matching metadata files found in source_folder.")

    groups = _group_by_run_date(files)
    outputs: List[Path] = []

    for (run, date), paths in groups.items():
        group_rows: List[Dict[str, object]] = []
        for p in paths:
            group_rows.extend(_read_sample_rows(p))

        if not group_rows:
            # If no rows, still emit an empty data section but preserve template
            out_path = out / f"{run}_{date}.xlsx"
            wb = load_workbook(tpl, data_only=False)
            wb.save(out_path)
            wb.close()
            outputs.append(out_path)
            continue

        out_path = _write_group(run, date, group_rows, tpl, out)
        outputs.append(out_path)

    return outputs